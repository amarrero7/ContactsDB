from __future__ import print_function
import pandas
import json
import requests
import time
import webbrowser
import os
import openpyxl
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

access_required = True
if access_required:
    access_token = None

    url = 'https://authz.constantcontact.com/oauth2/default/v1/device/authorize'
    data = {
        'client_id': "32fe8fe3-8097-4701-bebd-990b8a79a7e6",
        'scope': "contact_data account_update offline_access account_read",
    }

    autorization_request = requests.post(url, data=data)

    verification_uri_complete = autorization_request.json()['verification_uri_complete']
    device_code = autorization_request.json()['device_code']
    webbrowser.open(verification_uri_complete)

    verified = input("Have you autorized the app? (y/n): ")

    if verified == "y" or "Y":
        access_token_url = "https://authz.constantcontact.com/oauth2/default/v1/token"

        data = {
            'client_id': "32fe8fe3-8097-4701-bebd-990b8a79a7e6",
            'device_code': device_code,
            'grant_type': 'urn:ietf:params:oauth:grant-type:device_code',
        }

        headers = {
            'content_type': 'application/x-www-form-urlencoded',
        }

        access_token_request = requests.post(access_token_url, data=data, headers=headers)

        access_token = access_token_request.json()['access_token']

    if access_token_request.status_code == 200:
        print("Access token granted")
    else:
        print("Error: ", access_token_request.status_code)

def retrieve_contacts_lists(target_list_name):
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json',
    }

    url = 'https://api.cc.email/v3/contact_lists'

    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        print("Success")
    else:
        print("Error: ", response.status_code)

    for list in response.json()['lists']:
        if (list['name'] == target_list_name):
            print("Target List :" , list['name'], list['list_id'])

def add_contact(contact_data):
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json',
    }
    url = 'https://api.cc.email/v3/contacts/sign_up_form'
    response = requests.post(url, headers=headers, json=contact_data)
    if response.status_code == 200 or response.status_code == 201:
        print("Success -> ", response.status_code)
    else:
        print("Error: ", response.json())

cwd = os.getcwd()
contact_workbook = openpyxl.load_workbook(os.path.join(cwd, "contacts-for-verification.xlsx"))
sheet = contact_workbook["contacts-for-verification"]

for i in range(133, 138):

    email = sheet.cell(row=i, column=1).value
    firstname = sheet.cell(row=i, column=2).value
    lastname = sheet.cell(row=i, column=3).value
    company = sheet.cell(row=i, column=4).value
    role = sheet.cell(row=i, column=5).value
    country = sheet.cell(row=i, column=20).value

    contact_data = {
        "first_name": firstname,
        "last_name": lastname,
        "company_name": company,
        "job_title": role,
        "email_address": email,
        "street_address": { "country": country },
        "list_memberships": ["5ebf8b86-0567-11ee-95cf-fa163ec0786c"]
    }
    add_contact(contact_data)

# """Accesing the google spreadsheet that contains 
# all the contacts. Adding them to the Constant Contact.
# """
# # If modifying these scopes, delete the file token.json.
# SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

# # The ID and range of a sample spreadsheet.
# SAMPLE_SPREADSHEET_ID = '1IJBGEujMSVd9Ungi2XQQ7EpeI1mf_yz9Xxaed2Prg_E'
# creds = None

# # The file token.json stores the user's access and refresh tokens, and is
# # created automatically when the authorization flow completes for the first
# # time.
# if os.path.exists('token.json'):
#     creds = Credentials.from_authorized_user_file('token.json', SCOPES)
# # If there are no (valid) credentials available, let the user log in.
# if not creds or not creds.valid:
#     if creds and creds.expired and creds.refresh_token:
#         creds.refresh(Request())
#     else:
#         flow = InstalledAppFlow.from_client_secrets_file(
#             'credentials.json', SCOPES)
#         creds = flow.run_local_server(port=0)
#     # Save the credentials for the next run
#     with open('token.json', 'w') as token:
#         token.write(creds.to_json())

# try:
#     service = build('sheets', 'v4', credentials=creds)

#     # Call the Sheets API
#     sheet = service.spreadsheets()
#     result = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID,
#                                 range='B6:H').execute()
#     values = result.get('values', [])

#     if not values:
#         print('No data found.')

#     for row in values:
#         firstname = row[0]
#         lastname = row[1]
#         company = row[2]
#         role = row[3]
#         phone = row[5]
#         email = row[6]
#         country = row[4]
#         contact_data = {
#             "first_name": firstname,
#             "last_name": lastname,
#             "company_name": company,
#             "job_title": role,
#             "phone_number": phone,
#             "email_address": email,
#             "street_address": { "country": country },
#             "list_memberships": ["97de3134-03db-11ee-9a40-fa163e5fbd10"]
#         }
#         add_contact(contact_data)
# except HttpError as err:
#     print(err)


